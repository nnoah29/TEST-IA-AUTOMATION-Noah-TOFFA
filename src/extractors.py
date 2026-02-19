from typing import Optional, Tuple, Protocol
import os
import csv
from loguru import logger
from pypdf import PdfReader
import docx
import openpyxl

class Extractor(Protocol):
    def extract(self, file_path: str) -> Optional[str]:
        ...

class TextExtractor:
    def extract(self, file_path: str) -> Optional[str]:
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        except Exception as e:
            logger.error(f"Error extracting text from {file_path}: {e}")
            return None

class PDFExtractor:
    def extract(self, file_path: str) -> Optional[str]:
        try:
            reader = PdfReader(file_path)
            text = ""
            for page in reader.pages:
                text += page.extract_text() + "\n"
            return text
        except Exception as e:
            logger.error(f"Error extracting PDF from {file_path}: {e}")
            return None

class WordExtractor:
    def extract(self, file_path: str) -> Optional[str]:
        try:
            doc = docx.Document(file_path)
            text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
            return text
        except Exception as e:
            logger.error(f"Error extracting Word doc from {file_path}: {e}")
            return None

class ExcelExtractor:
    def extract(self, file_path: str) -> Optional[str]:
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            text = ""
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                text += f"Sheet: {sheet}\n"
                for row in ws.iter_rows(values_only=True):
                    row_text = " ".join([str(cell) for cell in row if cell is not None])
                    text += row_text + "\n"
            return text
        except Exception as e:
            logger.error(f"Error extracting Excel from {file_path}: {e}")
            return None

class CSVExtractor:
    def extract(self, file_path: str) -> Optional[str]:
        try:
            text = ""
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for row in reader:
                    text += " ".join(row) + "\n"
            return text
        except Exception as e:
            logger.error(f"Error extracting CSV from {file_path}: {e}")
            return None

def extract_content(file_path: str) -> Tuple[Optional[str], bool]:
    """
    Extracts content from a file.
    Returns a tuple (text_content, is_image).
    """
    ext = os.path.splitext(file_path)[1].lower()
    
    # Image handling
    if ext in ['.jpg', '.jpeg', '.png', '.webp']:
        return None, True

    extractor: Optional[Extractor] = None

    if ext in ['.txt', '.md', '.log']:
        extractor = TextExtractor()
    elif ext == '.pdf':
        extractor = PDFExtractor()
    elif ext == '.docx':
        extractor = WordExtractor()
    elif ext in ['.xlsx', '.xls']:
        extractor = ExcelExtractor()
    elif ext == '.csv':
        extractor = CSVExtractor()
    
    if extractor:
        return extractor.extract(file_path), False
    
    logger.warning(f"Unsupported file type: {ext}")
    return None, False